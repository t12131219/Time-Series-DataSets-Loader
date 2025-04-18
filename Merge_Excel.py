import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os  # 用于删除文件


# file_path = 'Result/UCR/ACSF1/ACSF1.xlsx'
# meta = pd.read_excel(file_path, sheet_name=0, header = None)
# file_name = meta[0][1]
# print('file_name:\n', file_name)
# data = pd.read_excel(file_path, sheet_name=1, header = None)
#
#
# # 转置数据
# data = np.transpose(data)
#
# # 创建新的DataFrame - 优化点：使用concat一次性合并列
# result = pd.concat([
#     pd.DataFrame({'file_name': [file_name] + ['']*(len(data)-1)}),
#     data
# ], axis=1)
#
# # 确保有6行数据（不足补空行）
# if len(result) < 6:
#     result = pd.concat([
#         result,
#         pd.DataFrame(np.nan, index=range(6-len(result)), columns=result.columns)
#     ], axis=0)
#
# # 保存到临时文件
# temp_path = 'temp_result.xlsx'
# result.to_excel(temp_path, index=False, header=False)
#
# # 用openpyxl处理合并单元格
# wb = load_workbook(temp_path)
# ws = wb.active
# ws.merge_cells('A1:A6')
# ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
#
# # 保存并清理
# final_path = 'final_result.xlsx'
# wb.save(final_path)
# if os.path.exists(temp_path):
#     os.remove(temp_path)
#
# print(f"处理完成！结果保存到: {final_path}")

# 文件路径列表

file_paths = [
    'Result/UCR/ACSF1/ACSF1.xlsx',
    'Result/UCR/Adiac/Adiac.xlsx',
    'Result/UCR/AllGestureWiimoteX/AllGestureWiimoteX.xlsx',
    'Result/UCR/AllGestureWiimoteY/AllGestureWiimoteY.xlsx',
    'Result/UCR/AllGestureWiimoteZ/AllGestureWiimoteZ.xlsx',
    'Result/UCR/ArrowHead/ArrowHead.xlsx',
    'Result/UCR/Beef/Beef.xlsx',
    'Result/UCR/BeetleFly/BeetleFly.xlsx',
    'Result/UCR/BirdChicken/BirdChicken.xlsx',
    'Result/UCR/BME/BME.xlsx',
    'Result/UCR/Car/Car.xlsx',
    'Result/UCR/CBF/CBF.xlsx',
    'Result/UCR/Chinatown/Chinatown.xlsx',
    'Result/UCR/ChlorineConcentration/ChlorineConcentration.xlsx',
    'Result/UCR/CinCECGTorso/CinCECGTorso.xlsx',
    'Result/UCR/Coffee/Coffee.xlsx',
    'Result/UCR/Computers/Computers.xlsx',
    'Result/UCR/CricketX/CricketX.xlsx',
    'Result/UCR/CricketY/CricketY.xlsx',
    'Result/UCR/CricketZ/CricketZ.xlsx',
    'Result/UCR/Crop/Crop.xlsx',
    'Result/UCR/DiatomSizeReduction/DiatomSizeReduction.xlsx',
    'Result/UCR/DistalPhalanxOutlineAgeGroup/DistalPhalanxOutlineAgeGroup.xlsx',
    'Result/UCR/DistalPhalanxOutlineCorrect/DistalPhalanxOutlineCorrect.xlsx',
    'Result/UCR/DistalPhalanxTW/DistalPhalanxTW.xlsx',
    'Result/UCR/DodgerLoopDay/DodgerLoopDay.xlsx',
    'Result/UCR/DodgerLoopGame/DodgerLoopGame.xlsx',
    'Result/UCR/DodgerLoopWeekend/DodgerLoopWeekend.xlsx',
    'Result/UCR/Earthquakes/Earthquakes.xlsx',
    'Result/UCR/ECG200/ECG200.xlsx',
    'Result/UCR/ECG5000/ECG5000.xlsx',
    'Result/UCR/ECGFiveDays/ECGFiveDays.xlsx',
    'Result/UCR/ElectricDevices/ElectricDevices.xlsx',
    'Result/UCR/EOGHorizontalSignal/EOGHorizontalSignal.xlsx',
    'Result/UCR/EOGVerticalSignal/EOGVerticalSignal.xlsx',
    'Result/UCR/EthanolLevel/EthanolLevel.xlsx',
    'Result/UCR/FaceAll/FaceAll.xlsx',
    'Result/UCR/FaceFour/FaceFour.xlsx',
    'Result/UCR/FacesUCR/FacesUCR.xlsx',
    'Result/UCR/FiftyWords/FiftyWords.xlsx',
    'Result/UCR/Fish/Fish.xlsx',
    'Result/UCR/FordA/FordA.xlsx',
    'Result/UCR/FordB/FordB.xlsx',
    'Result/UCR/FreezerRegularTrain/FreezerRegularTrain.xlsx',
    'Result/UCR/FreezerSmallTrain/FreezerSmallTrain.xlsx',
    'Result/UCR/Fungi/Fungi.xlsx',
    'Result/UCR/GestureMidAirD1/GestureMidAirD1.xlsx',
    'Result/UCR/GestureMidAirD2/GestureMidAirD2.xlsx',
    'Result/UCR/GestureMidAirD3/GestureMidAirD3.xlsx',
    'Result/UCR/GesturePebbleZ1/GesturePebbleZ1.xlsx',
    'Result/UCR/GesturePebbleZ2/GesturePebbleZ2.xlsx',
    'Result/UCR/GunPoint/GunPoint.xlsx',
    'Result/UCR/GunPointAgeSpan/GunPointAgeSpan.xlsx',
    'Result/UCR/GunPointMaleVersusFemale/GunPointMaleVersusFemale.xlsx',
    'Result/UCR/GunPointOldVersusYoung/GunPointOldVersusYoung.xlsx',
    'Result/UCR/Ham/Ham.xlsx',
    'Result/UCR/HandOutlines/HandOutlines.xlsx',
    'Result/UCR/Haptics/Haptics.xlsx',
    'Result/UCR/Herring/Herring.xlsx',
    'Result/UCR/HouseTwenty/HouseTwenty.xlsx',
    'Result/UCR/InlineSkate/InlineSkate.xlsx',
    'Result/UCR/InsectEPGRegularTrain/InsectEPGRegularTrain.xlsx',
    'Result/UCR/InsectEPGSmallTrain/InsectEPGSmallTrain.xlsx',
    'Result/UCR/InsectWingbeatSound/InsectWingbeatSound.xlsx',
    'Result/UCR/ItalyPowerDemand/ItalyPowerDemand.xlsx',
    'Result/UCR/LargeKitchenAppliances/LargeKitchenAppliances.xlsx',
    'Result/UCR/Lightning2/Lightning2.xlsx',
    'Result/UCR/Lightning7/Lightning7.xlsx',
    'Result/UCR/Mallat/Mallat.xlsx',
    'Result/UCR/Meat/Meat.xlsx',
    'Result/UCR/MedicalImages/MedicalImages.xlsx',
    'Result/UCR/MelbournePedestrian/MelbournePedestrian.xlsx',
    'Result/UCR/MiddlePhalanxOutlineAgeGroup/MiddlePhalanxOutlineAgeGroup.xlsx',
    'Result/UCR/MiddlePhalanxOutlineCorrect/MiddlePhalanxOutlineCorrect.xlsx',
    'Result/UCR/MiddlePhalanxTW/MiddlePhalanxTW.xlsx',
    'Result/UCR/MixedShapesRegularTrain/MixedShapesRegularTrain.xlsx',
    'Result/UCR/MixedShapesSmallTrain/MixedShapesSmallTrain.xlsx',
    'Result/UCR/MoteStrain/MoteStrain.xlsx',
    'Result/UCR/NonInvasiveFetalECGThorax1/NonInvasiveFetalECGThorax1.xlsx',
    'Result/UCR/NonInvasiveFetalECGThorax2/NonInvasiveFetalECGThorax2.xlsx',
    'Result/UCR/OliveOil/OliveOil.xlsx',
    'Result/UCR/OSULeaf/OSULeaf.xlsx',
    'Result/UCR/PhalangesOutlinesCorrect/PhalangesOutlinesCorrect.xlsx',
    'Result/UCR/Phoneme/Phoneme.xlsx',
    'Result/UCR/PickupGestureWiimoteZ/PickupGestureWiimoteZ.xlsx',
    'Result/UCR/PigAirwayPressure/PigAirwayPressure.xlsx',
    'Result/UCR/PigArtPressure/PigArtPressure.xlsx',
    'Result/UCR/PigCVP/PigCVP.xlsx',
    'Result/UCR/PLAID/PLAID.xlsx',
    'Result/UCR/Plane/Plane.xlsx',
    'Result/UCR/PowerCons/PowerCons.xlsx',
    'Result/UCR/ProximalPhalanxOutlineAgeGroup/ProximalPhalanxOutlineAgeGroup.xlsx',
    'Result/UCR/ProximalPhalanxOutlineCorrect/ProximalPhalanxOutlineCorrect.xlsx',
    'Result/UCR/ProximalPhalanxTW/ProximalPhalanxTW.xlsx',
    'Result/UCR/RefrigerationDevices/RefrigerationDevices.xlsx',
    'Result/UCR/Rock/Rock.xlsx',
    'Result/UCR/ScreenType/ScreenType.xlsx',
    'Result/UCR/SemgHandGenderCh2/SemgHandGenderCh2.xlsx',
    'Result/UCR/SemgHandMovementCh2/SemgHandMovementCh2.xlsx',
    'Result/UCR/SemgHandSubjectCh2/SemgHandSubjectCh2.xlsx',
    'Result/UCR/ShakeGestureWiimoteZ/ShakeGestureWiimoteZ.xlsx',
    'Result/UCR/ShapeletSim/ShapeletSim.xlsx',
    'Result/UCR/ShapesAll/ShapesAll.xlsx',
    'Result/UCR/SmallKitchenAppliances/SmallKitchenAppliances.xlsx',
    'Result/UCR/SmoothSubspace/SmoothSubspace.xlsx',
    'Result/UCR/SonyAIBORobotSurface1/SonyAIBORobotSurface1.xlsx',
    'Result/UCR/SonyAIBORobotSurface2/SonyAIBORobotSurface2.xlsx',
    'Result/UCR/StarLightCurves/StarLightCurves.xlsx',
    'Result/UCR/Strawberry/Strawberry.xlsx',
    'Result/UCR/SwedishLeaf/SwedishLeaf.xlsx',
    'Result/UCR/Symbols/Symbols.xlsx',
    'Result/UCR/SyntheticControl/SyntheticControl.xlsx',
    'Result/UCR/ToeSegmentation1/ToeSegmentation1.xlsx',
    'Result/UCR/ToeSegmentation2/ToeSegmentation2.xlsx',
    'Result/UCR/Trace/Trace.xlsx',
    'Result/UCR/TwoLeadECG/TwoLeadECG.xlsx',
    'Result/UCR/TwoPatterns/TwoPatterns.xlsx',
    'Result/UCR/UMD/UMD.xlsx',
    'Result/UCR/UWaveGestureLibraryAll/UWaveGestureLibraryAll.xlsx',
    'Result/UCR/UWaveGestureLibraryX/UWaveGestureLibraryX.xlsx',
    'Result/UCR/UWaveGestureLibraryY/UWaveGestureLibraryY.xlsx',
    'Result/UCR/UWaveGestureLibraryZ/UWaveGestureLibraryZ.xlsx',
    'Result/UCR/Wafer/Wafer.xlsx',
    'Result/UCR/Wine/Wine.xlsx',
    'Result/UCR/WordSynonyms/WordSynonyms.xlsx',
    'Result/UCR/Worms/Worms.xlsx',
    'Result/UCR/WormsTwoClass/WormsTwoClass.xlsx',
    'Result/UCR/Yoga/Yoga.xlsx',
]

file_paths = [
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/AppliancesEnergy/AppliancesEnergy_TEST/AppliancesEnergy_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/AppliancesEnergy/AppliancesEnergy_TRAIN/AppliancesEnergy_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/AustraliaRainfall/AustraliaRainfall_TEST/AustraliaRainfall_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/AustraliaRainfall/AustraliaRainfall_TRAIN/AustraliaRainfall_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/BeijingPM10Quality/BeijingPM10Quality_TEST/BeijingPM10Quality_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/BeijingPM10Quality/BeijingPM10Quality_TRAIN/BeijingPM10Quality_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/BeijingPM25Quality/BeijingPM25Quality_TEST/BeijingPM25Quality_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/BeijingPM25Quality/BeijingPM25Quality_TRAIN/BeijingPM25Quality_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/BenzeneConcentration/BenzeneConcentration_TEST/BenzeneConcentration_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/BenzeneConcentration/BenzeneConcentration_TRAIN/BenzeneConcentration_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/BIDMC32HR/BIDMC32HR_TEST/BIDMC32HR_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/BIDMC32HR/BIDMC32HR_TRAIN/BIDMC32HR_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/BIDMC32RR/BIDMC32RR_TEST/BIDMC32RR_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/BIDMC32RR/BIDMC32RR_TRAIN/BIDMC32RR_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/BIDMC32SpO2/BIDMC32SpO2_TEST/BIDMC32SpO2_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/BIDMC32SpO2/BIDMC32SpO2_TRAIN/BIDMC32SpO2_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/Covid3Month/Covid3Month_TEST/Covid3Month_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/Covid3Month/Covid3Month_TRAIN/Covid3Month_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/FloodModeling1/FloodModeling1_TEST/FloodModeling1_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/FloodModeling1/FloodModeling1_TRAIN/FloodModeling1_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/FloodModeling2/FloodModeling2_TEST/FloodModeling2_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/FloodModeling2/FloodModeling2_TRAIN/FloodModeling2_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/FloodModeling3/FloodModeling3_TEST/FloodModeling3_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/FloodModeling3/FloodModeling3_TRAIN/FloodModeling3_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/HouseholdPowerConsumption1/HouseholdPowerConsumption1_TEST/HouseholdPowerConsumption1_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/HouseholdPowerConsumption1/HouseholdPowerConsumption1_TRAIN/HouseholdPowerConsumption1_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/HouseholdPowerConsumption2/HouseholdPowerConsumption2_TEST/HouseholdPowerConsumption2_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/HouseholdPowerConsumption2/HouseholdPowerConsumption2_TRAIN/HouseholdPowerConsumption2_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/IEEEPPG/IEEEPPG_TEST/IEEEPPG_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/IEEEPPG/IEEEPPG_TRAIN/IEEEPPG_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/LiveFuelMoistureContent/LiveFuelMoistureContent_TEST/LiveFuelMoistureContent_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/LiveFuelMoistureContent/LiveFuelMoistureContent_TRAIN/LiveFuelMoistureContent_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/NewsHeadlineSentiment/NewsHeadlineSentiment_TEST/NewsHeadlineSentiment_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/NewsHeadlineSentiment/NewsHeadlineSentiment_TRAIN/NewsHeadlineSentiment_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/NewsTitleSentiment/NewsTitleSentiment_TEST/NewsTitleSentiment_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/NewsTitleSentiment/NewsTitleSentiment_TRAIN/NewsTitleSentiment_TRAIN.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/PPGDalia/PPGDalia_TEST/PPGDalia_TEST.ts.xlsx',
    'Result/datasets/TSER-Monash_UEA_UCR_Regression_Archive/PPGDalia/PPGDalia_TRAIN/PPGDalia_TRAIN.ts.xlsx'
]

# 初始化结果列表（避免DataFrame碎片化）
result_chunks = []

for file_path in file_paths:
    try:
        # 读取数据
        meta = pd.read_excel(file_path, sheet_name=0, header=None)
        file_name = meta[0][1]
        data = pd.read_excel(file_path, sheet_name=1, header=None)

        print(f"————————处理文件 {file_path} ————————")

        # 转置数据并确保不超过6行
        data = np.transpose(data).head(6)

        # 创建当前文件的DataFrame（优化点：一次性构建）
        current_df = pd.DataFrame({
            'file_name': [file_name] + [''] * (len(data) - 1),
            ** {f'col_{i}': data[i] for i in data.columns}
        })

        # 填充到6行（优化点：使用concat替代逐行添加）
        if len(current_df) < 6:
            filler = pd.DataFrame(
                [[''] + [np.nan] * len(data.columns)] * (6 - len(current_df)),
                columns=current_df.columns
            )
            current_df = pd.concat([current_df, filler], ignore_index=True)

        # 添加到结果列表
        result_chunks.append(current_df.copy())  # 使用copy避免碎片化

    except Exception as e:
        print(f"处理文件 {file_path} 时出错: {str(e)}")
        continue

# 一次性合并所有结果（优化点：避免迭代concat）
if result_chunks:
    final_result = pd.concat(result_chunks, ignore_index=True)
else:
    final_result = pd.DataFrame()

# 保存到临时文件
temp_path = 'temp_combined.xlsx'
final_result.to_excel(temp_path, index=False, header=False)

# 用openpyxl处理合并单元格
wb = load_workbook(temp_path)
ws = wb.active

# 为每个数据集合并A1:A6单元格
for i in range(0, len(final_result), 6):
    start_row = i + 1
    end_row = start_row + 5
    ws.merge_cells(f'A{start_row}:A{end_row}')
    ws[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')

# 保存最终结果
final_path = 'final_result.xlsx'
wb.save(final_path)

# 删除临时文件
if os.path.exists(temp_path):
    os.remove(temp_path)

print(f"处理完成！合并结果保存到: {final_path}")


'''
# 示例用法
file_paths = [
    'Result/UCR/ACSF1/ACSF1.xlsx',
    'Result/UCR/Adiac/Adiac.xlsx',
    'Result/UCR/AllGestureWiimoteX/AllGestureWiimoteX.xlsx',
    'Result/UCR/AllGestureWiimoteY/AllGestureWiimoteY.xlsx',
    'Result/UCR/AllGestureWiimoteZ/AllGestureWiimoteZ.xlsx',
    'Result/UCR/ArrowHead/ArrowHead.xlsx',
    'Result/UCR/Beef/Beef.xlsx',
    'Result/UCR/BeetleFly/BeetleFly.xlsx',
    'Result/UCR/BirdChicken/BirdChicken.xlsx',
    'Result/UCR/BME/BME.xlsx',
    'Result/UCR/Car/Car.xlsx',
    'Result/UCR/CBF/CBF.xlsx',
    'Result/UCR/Chinatown/Chinatown.xlsx',
    'Result/UCR/ChlorineConcentration/ChlorineConcentration.xlsx',
    'Result/UCR/CinCECGTorso/CinCECGTorso.xlsx',
    'Result/UCR/Coffee/Coffee.xlsx',
    'Result/UCR/Computers/Computers.xlsx',
    'Result/UCR/CricketX/CricketX.xlsx',
    'Result/UCR/CricketY/CricketY.xlsx',
    'Result/UCR/CricketZ/CricketZ.xlsx',
    'Result/UCR/Crop/Crop.xlsx',
    'Result/UCR/DiatomSizeReduction/DiatomSizeReduction.xlsx',
    'Result/UCR/DistalPhalanxOutlineAgeGroup/DistalPhalanxOutlineAgeGroup.xlsx',
    'Result/UCR/DistalPhalanxOutlineCorrect/DistalPhalanxOutlineCorrect.xlsx',
    'Result/UCR/DistalPhalanxTW/DistalPhalanxTW.xlsx',
    'Result/UCR/DodgerLoopDay/DodgerLoopDay.xlsx',
    'Result/UCR/DodgerLoopGame/DodgerLoopGame.xlsx',
    'Result/UCR/DodgerLoopWeekend/DodgerLoopWeekend.xlsx',
    'Result/UCR/Earthquakes/Earthquakes.xlsx',
    'Result/UCR/ECG200/ECG200.xlsx',
    'Result/UCR/ECG5000/ECG5000.xlsx',
    'Result/UCR/ECGFiveDays/ECGFiveDays.xlsx',
    'Result/UCR/ElectricDevices/ElectricDevices.xlsx',
    'Result/UCR/EOGHorizontalSignal/EOGHorizontalSignal.xlsx',
    'Result/UCR/EOGVerticalSignal/EOGVerticalSignal.xlsx',
    'Result/UCR/EthanolLevel/EthanolLevel.xlsx',
    'Result/UCR/FaceAll/FaceAll.xlsx',
    'Result/UCR/FaceFour/FaceFour.xlsx',
    'Result/UCR/FacesUCR/FacesUCR.xlsx',
    'Result/UCR/FiftyWords/FiftyWords.xlsx',
    'Result/UCR/Fish/Fish.xlsx',
    'Result/UCR/FordA/FordA.xlsx',
    'Result/UCR/FordB/FordB.xlsx',
    'Result/UCR/FreezerRegularTrain/FreezerRegularTrain.xlsx',
    'Result/UCR/FreezerSmallTrain/FreezerSmallTrain.xlsx',
    'Result/UCR/Fungi/Fungi.xlsx',
    'Result/UCR/GestureMidAirD1/GestureMidAirD1.xlsx',
    'Result/UCR/GestureMidAirD2/GestureMidAirD2.xlsx',
    'Result/UCR/GestureMidAirD3/GestureMidAirD3.xlsx',
    'Result/UCR/GesturePebbleZ1/GesturePebbleZ1.xlsx',
    'Result/UCR/GesturePebbleZ2/GesturePebbleZ2.xlsx',
    'Result/UCR/GunPoint/GunPoint.xlsx',
    'Result/UCR/GunPointAgeSpan/GunPointAgeSpan.xlsx',
    'Result/UCR/GunPointMaleVersusFemale/GunPointMaleVersusFemale.xlsx',
    'Result/UCR/GunPointOldVersusYoung/GunPointOldVersusYoung.xlsx',
    'Result/UCR/Ham/Ham.xlsx',
    'Result/UCR/HandOutlines/HandOutlines.xlsx',
    'Result/UCR/Haptics/Haptics.xlsx',
    'Result/UCR/Herring/Herring.xlsx',
    'Result/UCR/HouseTwenty/HouseTwenty.xlsx',
    'Result/UCR/InlineSkate/InlineSkate.xlsx',
    'Result/UCR/InsectEPGRegularTrain/InsectEPGRegularTrain.xlsx',
    'Result/UCR/InsectEPGSmallTrain/InsectEPGSmallTrain.xlsx',
    'Result/UCR/InsectWingbeatSound/InsectWingbeatSound.xlsx',
    'Result/UCR/ItalyPowerDemand/ItalyPowerDemand.xlsx',
    'Result/UCR/LargeKitchenAppliances/LargeKitchenAppliances.xlsx',
    'Result/UCR/Lightning2/Lightning2.xlsx',
    'Result/UCR/Lightning7/Lightning7.xlsx',
    'Result/UCR/Mallat/Mallat.xlsx',
    'Result/UCR/Meat/Meat.xlsx',
    'Result/UCR/MedicalImages/MedicalImages.xlsx',
    'Result/UCR/MelbournePedestrian/MelbournePedestrian.xlsx',
    'Result/UCR/MiddlePhalanxOutlineAgeGroup/MiddlePhalanxOutlineAgeGroup.xlsx',
    'Result/UCR/MiddlePhalanxOutlineCorrect/MiddlePhalanxOutlineCorrect.xlsx',
    'Result/UCR/MiddlePhalanxTW/MiddlePhalanxTW.xlsx',
    'Result/UCR/MixedShapesRegularTrain/MixedShapesRegularTrain.xlsx',
    'Result/UCR/MixedShapesSmallTrain/MixedShapesSmallTrain.xlsx',
    'Result/UCR/MoteStrain/MoteStrain.xlsx',
    'Result/UCR/NonInvasiveFetalECGThorax1/NonInvasiveFetalECGThorax1.xlsx',
    'Result/UCR/NonInvasiveFetalECGThorax2/NonInvasiveFetalECGThorax2.xlsx',
    'Result/UCR/OliveOil/OliveOil.xlsx',
    'Result/UCR/OSULeaf/OSULeaf.xlsx',
    'Result/UCR/PhalangesOutlinesCorrect/PhalangesOutlinesCorrect.xlsx',
    'Result/UCR/Phoneme/Phoneme.xlsx',
    'Result/UCR/PickupGestureWiimoteZ/PickupGestureWiimoteZ.xlsx',
    'Result/UCR/PigAirwayPressure/PigAirwayPressure.xlsx',
    'Result/UCR/PigArtPressure/PigArtPressure.xlsx',
    'Result/UCR/PigCVP/PigCVP.xlsx',
    'Result/UCR/PLAID/PLAID.xlsx',
    'Result/UCR/Plane/Plane.xlsx',
    'Result/UCR/PowerCons/PowerCons.xlsx',
    'Result/UCR/ProximalPhalanxOutlineAgeGroup/ProximalPhalanxOutlineAgeGroup.xlsx',
    'Result/UCR/ProximalPhalanxOutlineCorrect/ProximalPhalanxOutlineCorrect.xlsx',
    'Result/UCR/ProximalPhalanxTW/ProximalPhalanxTW.xlsx',
    'Result/UCR/RefrigerationDevices/RefrigerationDevices.xlsx',
    'Result/UCR/Rock/Rock.xlsx',
    'Result/UCR/ScreenType/ScreenType.xlsx',
    'Result/UCR/SemgHandGenderCh2/SemgHandGenderCh2.xlsx',
    'Result/UCR/SemgHandMovementCh2/SemgHandMovementCh2.xlsx',
    'Result/UCR/SemgHandSubjectCh2/SemgHandSubjectCh2.xlsx',
    'Result/UCR/ShakeGestureWiimoteZ/ShakeGestureWiimoteZ.xlsx',
    'Result/UCR/ShapeletSim/ShapeletSim.xlsx',
    'Result/UCR/ShapesAll/ShapesAll.xlsx',
    'Result/UCR/SmallKitchenAppliances/SmallKitchenAppliances.xlsx',
    'Result/UCR/SmoothSubspace/SmoothSubspace.xlsx',
    'Result/UCR/SonyAIBORobotSurface1/SonyAIBORobotSurface1.xlsx',
    'Result/UCR/SonyAIBORobotSurface2/SonyAIBORobotSurface2.xlsx',
    'Result/UCR/StarLightCurves/StarLightCurves.xlsx',
    'Result/UCR/Strawberry/Strawberry.xlsx',
    'Result/UCR/SwedishLeaf/SwedishLeaf.xlsx',
    'Result/UCR/Symbols/Symbols.xlsx',
    'Result/UCR/SyntheticControl/SyntheticControl.xlsx',
    'Result/UCR/ToeSegmentation1/ToeSegmentation1.xlsx',
    'Result/UCR/ToeSegmentation2/ToeSegmentation2.xlsx',
    'Result/UCR/Trace/Trace.xlsx',
    'Result/UCR/TwoLeadECG/TwoLeadECG.xlsx',
    'Result/UCR/TwoPatterns/TwoPatterns.xlsx',
    'Result/UCR/UMD/UMD.xlsx',
    'Result/UCR/UWaveGestureLibraryAll/UWaveGestureLibraryAll.xlsx',
    'Result/UCR/UWaveGestureLibraryX/UWaveGestureLibraryX.xlsx',
    'Result/UCR/UWaveGestureLibraryY/UWaveGestureLibraryY.xlsx',
    'Result/UCR/UWaveGestureLibraryZ/UWaveGestureLibraryZ.xlsx',
    'Result/UCR/Wafer/Wafer.xlsx',
    'Result/UCR/Wine/Wine.xlsx',
    'Result/UCR/WordSynonyms/WordSynonyms.xlsx',
    'Result/UCR/Worms/Worms.xlsx',
    'Result/UCR/WormsTwoClass/WormsTwoClass.xlsx',
    'Result/UCR/Yoga/Yoga.xlsx',
]
'''

# # 转置数据
# data = np.transpose(data)
#
# # 创建新的DataFrame，预留6行空间
# result = pd.DataFrame(index=range(6), columns=data.columns)
#
# # 将file_name写入A1（后续合并后会跨A1:A6）
# result.iloc[0, 0] = file_name
#
# # 将数据从第7行开始拼接（跳过前6行）
# result = pd.concat([result, data], ignore_index=True)
#
# # 保存到临时文件
# temp_path = 'temp_result.xlsx'
# result.to_excel(temp_path, index=False, header=False)
#
# # 用openpyxl合并单元格并调整格式
# wb = load_workbook(temp_path)
# ws = wb.active
#
# # 纵向合并 A1:A6
# ws.merge_cells('A1:A6')
#
# # 设置垂直居中
# ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
#
# # 保存最终结果
# final_path = 'final_result.xlsx'
# wb.save(final_path)
#
# # 删除临时文件
# if os.path.exists(temp_path):
#     os.remove(temp_path)
#     print(f"临时文件 {temp_path} 已删除")
# else:
#     print(f"临时文件 {temp_path} 不存在")
#
# print(f"合并完成！结果保存到: {final_path}")

# # 显示数据的前几行
# print(meta.head())
#
# # 显示数据的前几行
# print(data.head())
#
#
# data = np.transpose(data)
#
# # 显示数据的前几行
# print(data.head())
#
#
# columns = data.columns